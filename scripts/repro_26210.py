#!/usr/bin/env python3
"""26210期复现: P5 开奖09451, 存档预测千9 0/10(万0 1/10/百4 2/10/十5 1/10/个1 1/10)
用截断数据(至26209)跑全量predict, 复现存档Top10, 追踪千位9丢失原因"""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

import openpyxl, tempfile
from p5_fusion_complete import Pick5FusionComplete, VERSION

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    '..', 'assets', 'data', '排列5历史数据.xlsx')
wb = openpyxl.load_workbook(DATA, read_only=True)
rows = list(wb.active.iter_rows(values_only=True))
rows = rows[:-1]  # 截断去掉26210
print(f"[repro-26210] 截断数据至 {rows[-1][0]} (共{len(rows)-1}期), 版本={VERSION}")

tmp = os.path.join(tempfile.gettempdir(), 'p5_trunc_26209.xlsx')
wb2 = openpyxl.Workbook(); ws2 = wb2.active
for r in rows: ws2.append(r)
wb2.save(tmp)

model = Pick5FusionComplete(data_path=tmp)
result = model.predict()
print("\n=== 复现Top10 ===")
for i, c in enumerate(result['bets'][:10]):
    print(f"  {i+1}. {''.join(map(str,c['digits']))} score={c['final_score']:.2f}")

top10 = [''.join(map(str,c['digits'])) for c in result['bets'][:10]]
print("\n存档预测: ['04885','15890','68081','65553','71225','96526','23070','96493','24117','01262']")
print("复现结果:", top10)
print("一致:", top10 == ['04885','15890','68081','65553','71225','96526','23070','96493','24117','01262'])

draw = (0, 9, 4, 5, 1)
for pos, name in [(0,'万'),(1,'千'),(2,'百'),(3,'十'),(4,'个')]:
    cnt = sum(1 for c in result['bets'][:10] if c['digits'][pos]==draw[pos])
    print(f"  {name}{draw[pos]}: {cnt}/10")
